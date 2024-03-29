import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment
from copy import copy
from flask import Response
import langid


def GetStructureNodeById(nodeId, nodes):
	for node in nodes:
		if(node['id'] == nodeId):
			return node
	return None

def GetStructureNodeChildrenById(parentNodeId, nodes):
	children = []
	for node in nodes:
		if(node['parent'] == parentNodeId):
			children.append(node)
	return children

def GetFailureProperty(key, failure):
	if('properties' not in failure):
		return ''

	result = []
	for pro in failure['properties']:
		if(pro['key'] == key):
			result.append(pro['value'])

	return '\n'.join(result)

def ExtendTable(sheet, lastRow):
	templateTableLastRow = 21
	templateTableStartColumn = 2
	templateTableEndColumn = 33

	if(lastRow > templateTableLastRow):
		for row in range(templateTableLastRow + 1, lastRow):
			for col in range(templateTableStartColumn, templateTableEndColumn):
				sheet.cell(row, col).border = copy(sheet.cell(templateTableLastRow  - 1, col).border) 
				sheet.cell(row, col).font = copy(sheet.cell(templateTableLastRow  - 1, col).font) 
				sheet.cell(row, col).alignment = Alignment(horizontal='center')
		# format last row
		for col in range(templateTableStartColumn, templateTableEndColumn):
			sheet.cell(lastRow, col).border = copy(sheet.cell(templateTableLastRow, col).border) 
			sheet.cell(lastRow, col).font = copy(sheet.cell(templateTableLastRow, col).font) 
			sheet.cell(lastRow, col).alignment = Alignment(horizontal='center')
			sheet.cell(templateTableLastRow, col).border = copy(sheet.cell(templateTableLastRow - 1, col).border)

def ResizeColumnWidth(sheet, lastRow):
	for col in sheet.columns:
	    max_length = 0
	    column = col[0].column_letter 
	    for cell in col[11:lastRow]:
	        if cell.coordinate in sheet.merged_cells:
	        	continue
	        try:
	        	cellWidth = getCellWidth(cell.value)
	        	if cellWidth > max_length:
	        		max_length = cellWidth
	        except:
	            pass
	    adjusted_width = (max_length + 2) * 1.2
	    sheet.column_dimensions[column].width = adjusted_width

def as_text(value):
    if value is None:
        return "####"
    return str(value)

''' performance is not accpected by check every charactor
def getCellWidth(value):
	text = as_text(value)
	getLenth = lambda x: 1 if(langid.classify(x)[0] == 'en') else 2
	return sum([getLenth(x) for x in text])
'''
'''
def getCellWidth(value):
	text = as_text(value)
	if(langid.classify(text)[0] == 'en'):
		return len(text)
	else:
		return 2 * len(text)
'''
def getCellWidth(value):
	text = as_text(value)
	return 2 * len(text)

def PutData(row, column, structureNode, workBookSheet, nodeType):
	lastRowIndex = row
	workBookSheet.cell(lastRowIndex, column).value = structureNode['name']
	if(len(structureNode['FunctionSet']) == 0):
		lastRowIndex = lastRowIndex + 1
	else:
		for function in structureNode['FunctionSet']:
			workBookSheet.cell(lastRowIndex, column + 3).value = function['name']
			if(len(function['FailureSet']) == 0):
				lastRowIndex = lastRowIndex + 1
			else:
				for failure in function['FailureSet']:
					if(nodeType == 'parent'):
						workBookSheet.cell(lastRowIndex, column + 6).value = failure['name']
						workBookSheet.cell(lastRowIndex, column + 7).value = failure['sValue']
					else:
						workBookSheet.cell(lastRowIndex, column + 7).value = failure['name']
					if(nodeType == 'curent'):
						workBookSheet.cell(lastRowIndex, column + 9).value = GetFailureProperty(1, failure)
						workBookSheet.cell(lastRowIndex, column + 10).value = failure['oValue']
						workBookSheet.cell(lastRowIndex, column + 11).value = GetFailureProperty(2, failure)
						workBookSheet.cell(lastRowIndex, column + 12).value = failure['dValue']
					lastRowIndex = lastRowIndex + 1

	return lastRowIndex

def ExportDataToTemplate(structureNodeId, structureNodes):
	wb = openpyxl.load_workbook(filename='template01.xlsx')
	sheet = wb.active

	sNode = GetStructureNodeById(structureNodeId, structureNodes)
	sNodeParent = GetStructureNodeById(sNode['parent'], structureNodes)
	sNodeChildren = GetStructureNodeChildrenById(structureNodeId, structureNodes)

	if not (sNodeParent is None):
		PutData(12, 4, sNodeParent, sheet, 'parent')
	if not(sNode is None):
		PutData(12, 5, sNode, sheet, 'curent')
	lastRowIndex = 12
	for childNode in sNodeChildren:
		lastRowIndex = PutData(lastRowIndex, 6, childNode, sheet, 'child')

	# extend table
	lastRow = lastRowIndex - 1
	ExtendTable(sheet, lastRow)

	ResizeColumnWidth(sheet, lastRow)

	response = Response(save_virtual_workbook(wb), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={"Content-Disposition":"attachment; filename=export.xlsx"})
	return response
	#wb.save(filename='export.xlsx')
